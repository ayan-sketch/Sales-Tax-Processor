#!/usr/bin/env python3
"""Generate sample Excel templates for IRIS return automation."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SHEET_COLORS = {"Individual Return": "4472C4", "WHT Statement": "ED7D31", "Company Return": "70AD47"}

FIELDS = {
    "Individual Return": [
        ("ntn", "Taxpayer NTN (e.g., 1234567-8)"),
        ("name", "Full name as on IRIS"),
        ("cnic", "CNIC number (13 digits)"),
        ("email", "Email address"),
        ("phone", "Phone/mobile number"),
        ("address", "Postal address"),
        ("tax_year", "Tax year (e.g., 2026)"),
        ("salary_income", "Total salary income"),
        ("salary_tax_deducted", "Tax deducted on salary"),
        ("business_income", "Business/profession income"),
        ("business_cost", "Business cost of goods sold"),
        ("business_tax_deducted", "Tax deducted on business income"),
        ("property_income", "Income from property (rent)"),
        ("property_tax_deducted", "Tax deducted on property income"),
        ("capital_gains", "Capital gains total"),
        ("capital_gains_tax", "Tax on capital gains"),
        ("other_income", "Income from other sources"),
        ("other_income_tax", "Tax deducted on other income"),
        ("donations", "Donations/charitable contributions"),
        ("investment_deduction", "Investment deductions"),
        ("education_expense", "Education expenses"),
        ("medical_expense", "Medical expenses"),
        ("tax_credits", "Tax credits claimed"),
        ("wealth_property", "Immovable property value"),
        ("wealth_movable", "Movable assets value"),
        ("wealth_cash", "Cash & bank balances"),
        ("wealth_investments", "Investments value"),
        ("wealth_liabilities", "Total liabilities"),
        ("taxable_income", "Total taxable income"),
        ("tax_payable", "Total tax payable"),
    ],
}


def create_template(name: str, fields: list[tuple[str, str]], path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name

    headers = ["Field", "Description", "Value"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20

    for r, (field, desc) in enumerate(fields, 2):
        ws.cell(row=r, column=1, value=field).font = Font(bold=True)
        ws.cell(row=r, column=2, value=desc)
        ws.cell(row=r, column=3, value="")

    wb.save(path)
    print(f"Created: {path}")


if __name__ == "__main__":
    import os
    out = os.path.dirname(__file__)
    for name, fields in FIELDS.items():
        path = os.path.join(out, name.lower().replace(" ", "_") + ".xlsx")
        create_template(name, fields, path)

    print("\nTemplates created. Fill column C with data and run the automation.")
